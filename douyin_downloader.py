#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
抖音视频无水印下载器 v1.2.4
支持单个视频下载、Excel 批量下载、关键词检测（语音+字幕OCR）
"""

import argparse
import asyncio
import json
import os
import re
import sys
import subprocess
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Optional, List, Dict
from urllib.parse import urlparse

# 强制刷新输出，确保实时显示日志
import functools
print = functools.partial(print, flush=True)

try:
    from playwright.async_api import async_playwright, Browser, Page
except ImportError:
    print("请先安装 playwright: pip install playwright")
    print("然后运行: playwright install chromium")
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    openpyxl = None  # Excel 功能可选

# 语音识别
try:
    from faster_whisper import WhisperModel
    WHISPER_AVAILABLE = True
except ImportError:
    WHISPER_AVAILABLE = False

# OCR 字幕识别
try:
    import easyocr
    import cv2
    import numpy as np
    OCR_AVAILABLE = True
except ImportError:
    OCR_AVAILABLE = False

# 简繁转换
try:
    import opencc
    OPENCC_AVAILABLE = True
except ImportError:
    OPENCC_AVAILABLE = False

# 全局变量：语音识别模型大小
MODEL_SIZE = "base"


class DetectionReporter:
    """检测结果报告器 - 将结果输出到文件"""

    def __init__(self, output_dir: Path, excel_mode: bool = False):
        self.output_dir = output_dir
        self.excel_mode = excel_mode
        self.report_file = output_dir / f"detection_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        self.results = []

    def add_result(self, video_name: str, url: str, speech_keywords: Dict, ocr_keywords: Dict, full_speech: str = '', full_ocr: str = ''):
        """添加检测结果"""
        result = {
            'video': video_name,
            'url': url,
            'speech_keywords': speech_keywords,
            'ocr_keywords': ocr_keywords,
            'full_speech': full_speech,
            'full_ocr': full_ocr,
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        self.results.append(result)
        self._write_result(result)

    def _write_result(self, result: Dict):
        """写入单个结果到文件"""
        with open(self.report_file, 'a', encoding='utf-8') as f:
            f.write(f"\n{'='*60}\n")
            f.write(f"视频: {result['video']}\n")
            f.write(f"链接: {result['url']}\n")
            f.write(f"时间: {result['timestamp']}\n")
            f.write(f"{'='*60}\n")

            # 语音检测结果
            if result['speech_keywords']:
                f.write("\n【语音检测】发现关键词:\n")
                for kw, occurrences in result['speech_keywords'].items():
                    f.write(f"  \"{kw}\" 出现 {len(occurrences)} 次\n")
                    for occ in occurrences:
                        f.write(f"    - {occ['timestamp']}: {occ['text']}\n")
            else:
                f.write("\n【语音检测】未发现关键词\n")

            # OCR 检测结果
            if result['ocr_keywords']:
                f.write("\n【字幕OCR检测】发现关键词:\n")
                for kw, occurrences in result['ocr_keywords'].items():
                    f.write(f"  \"{kw}\" 出现 {len(occurrences)} 次\n")
                    for occ in occurrences:
                        f.write(f"    - {occ['timestamp']}: {occ['text']}\n")
            else:
                f.write("\n【字幕OCR检测】未发现关键词\n")

            f.write("\n")

    def write_summary(self):
        """写入汇总"""
        with open(self.report_file, 'a', encoding='utf-8') as f:
            f.write(f"\n{'#'*60}\n")
            f.write("检测汇总\n")
            f.write(f"{'#'*60}\n")
            f.write(f"共检测 {len(self.results)} 个视频\n")

            # 统计有关键词的视频
            videos_with_keywords = []
            for r in self.results:
                if r['speech_keywords'] or r['ocr_keywords']:
                    videos_with_keywords.append(r['video'])

            if videos_with_keywords:
                f.write(f"\n发现关键词的视频 ({len(videos_with_keywords)} 个):\n")
                for v in videos_with_keywords:
                    f.write(f"  - {v}\n")
            else:
                f.write("\n所有视频均未发现关键词\n")

        print(f"\n[OK] 检测报告已保存: {self.report_file}")


class KeywordDetector:
    """关键词检测器 - 语音识别 + 字幕OCR"""

    def __init__(self, model_size: str = "base", output_dir: Optional[Path] = None):
        # 使用全局 MODEL_SIZE 变量
        self.whisper_model = None
        self.output_dir = output_dir or Path(".")
        self.ocr_reader = None
        self._converter = None

    def _to_traditional(self, text: str) -> str:
        """将简体中文转换为繁体中文"""
        try:
            import opencc
            if self._converter is None:
                self._converter = opencc.OpenCC('s2t')  # 简→繁
            return self._converter.convert(text)
        except ImportError:
            return text
        except Exception:
            return text

    def _load_whisper(self):
        """懒加载语音识别模型"""
        if self.whisper_model is None:
            if not WHISPER_AVAILABLE:
                return None
            print(f"  正在加载语音识别模型 ({MODEL_SIZE})...")
            self.whisper_model = WhisperModel(MODEL_SIZE, device="cpu", compute_type="int8")
        return self.whisper_model

    def _load_ocr(self):
        """懒加载 OCR 模型"""
        if self.ocr_reader is None:
            if not OCR_AVAILABLE:
                return None
            print(f"  正在加载 OCR 模型...")
            self.ocr_reader = easyocr.Reader(['ch_sim', 'en'], gpu=False)
        return self.ocr_reader

    def detect_speech(self, video_path: Path, keywords: List[str]) -> Dict:
        """检测语音中的关键词"""
        result = {'keywords_found': {}, 'transcript': '', 'success': False}

        try:
            model = self._load_whisper()
            if model is None:
                result['error'] = '语音识别模型不可用'
                return result

            print(f"  正在识别语音...")
            # 使用 zh 语言代码，不启用 VAD
            segments, info = model.transcribe(
                str(video_path),
                language="zh",
                beam_size=5,
                best_of=5
            )

            # 繁体转简体映射表（常用字）
            traditional_to_simplified = {
                '臺': '台', '麵': '面', '後': '后', '裡': '里', '鬆': '松', '鬥': '斗',
                '餘': '余', '穀': '谷', '醜': '丑', '捲': '卷', '鬍': '胡', '髮': '发',
                '後': '后', '徵': '征', '醜': '丑', '颱': '台', '夥': '伙', '彆': '别扭',
                '幹': '干', '醃': '腌', '嚮': '向', '薑': '姜', '鬆': '松', '鬱': '郁',
                '唄': '呗', '嘩': '哗', '嘩': '哗', '囉': '啰', '颱': '台', '籲': '吁',
                '鑊': '锅', '嚮': '向', '徵': '征', '醜': '丑', '捲': '卷', '鬍': '胡',
                '髮': '发', '後': '后', '徵': '征', '醜': '丑', '颱': '台', '夥': '伙',
                '彆': '别扭', '幹': '干', '醃': '腌', '嚮': '向', '薑': '姜', '鬆': '松',
                '鬱': '郁', '唄': '呗', '嘩': '哗', '囉': '啰', '籲': '吁', '鑊': '锅'
            }

            def traditional_to_simplified(text):
                """将繁体中文转换为简体中文"""
                result_text = text
                for traditional, simplified in traditional_to_simplified.items():
                    result_text = result_text.replace(traditional, simplified)
                return result_text

            transcript_parts = []
            all_text = []  # 收集所有识别到的文字

            for segment in segments:
                text = segment.text.strip()
                if text:
                    timestamp = f"{int(segment.start // 60):02d}:{int(segment.start % 60):02d}"
                    transcript_parts.append(f"[{timestamp}] {text}")
                    all_text.append(text.lower())  # 转小写用于匹配

                    # 检测关键词（大小写不敏感，将关键字转为繁体后匹配）
                    for keyword in keywords:
                        keyword_traditional = self._to_traditional(keyword)
                        if keyword_traditional.lower() in text.lower():
                            if keyword not in result['keywords_found']:
                                result['keywords_found'][keyword] = []
                            result['keywords_found'][keyword].append({
                                'timestamp': timestamp,
                                'text': text
                            })

            result['transcript'] = '\n'.join(transcript_parts)
            result['all_text'] = ' '.join(all_text)  # 保存所有文字
            result['success'] = True

        except Exception as e:
            result['error'] = str(e)

        # 保存到临时文件以便查看
        transcript_file = self.output_dir / "debug_speech.txt"
        with open(transcript_file, 'w', encoding='utf-8-sig') as f:  # 使用 utf-8-sig 避免乱码
            f.write(result.get('transcript', '无内容'))
        print(f"\n  [语音识别结果已保存到：{transcript_file}]\n")

        return result

    def detect_ocr(self, video_path: Path, keywords: List[str], interval: int = 1) -> Dict:
        """
        检测字幕中的关键词（OCR）

        Args:
            video_path: 视频路径
            keywords: 关键词列表
            interval: 抽帧间隔（秒）
        """
        result = {'keywords_found': {}, 'success': False}

        try:
            reader = self._load_ocr()
            if reader is None:
                result['error'] = 'OCR 模型不可用，请安装: pip install easyocr opencv-python'
                return result

            print(f"  正在识别字幕 (OCR)...")

            # 打开视频
            cap = cv2.VideoCapture(str(video_path))
            if not cap.isOpened():
                result['error'] = '无法打开视频文件'
                return result

            fps = cap.get(cv2.CAP_PROP_FPS)
            total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
            total_duration = total_frames / fps if fps > 0 else 0
            frame_interval = int(fps * 0.3)  # 每隔 0.3 秒抽一帧，更密集
            total_checks = total_frames // frame_interval if frame_interval > 0 else 0

            print(f"  视频时长: {int(total_duration//60)}分{int(total_duration%60)}秒, 需检测约 {total_checks} 帧")

            frame_count = 0
            check_count = 0
            detected_texts = []

            while True:
                ret, frame = cap.read()
                if not ret:
                    break

                if frame_count % frame_interval == 0:
                    check_count += 1
                    # 显示进度
                    if check_count % 10 == 0 or check_count == 1:
                        progress = check_count / total_checks * 100 if total_checks > 0 else 0
                        timestamp_sec = frame_count / fps
                        print(f"    OCR进度: {check_count}/{total_checks} ({progress:.0f}%) - {int(timestamp_sec//60):02d}:{int(timestamp_sec%60):02d}")

                    # OCR 识别
                    ocr_results = reader.readtext(frame)

                    # 计算时间戳
                    timestamp_sec = frame_count / fps
                    timestamp = f"{int(timestamp_sec // 60):02d}:{int(timestamp_sec % 60):02d}"

                    # 合并识别到的文字
                    frame_text = ' '.join([item[1] for item in ocr_results if item[2] > 0.3])  # 置信度 > 0.3

                    if frame_text.strip():
                        detected_texts.append(f"[{timestamp}] {frame_text}")

                        # 检测关键词
                        for keyword in keywords:
                            if keyword in frame_text:
                                if keyword not in result['keywords_found']:
                                    result['keywords_found'][keyword] = []
                                result['keywords_found'][keyword].append({
                                    'timestamp': timestamp,
                                    'text': frame_text[:100]
                                })

                frame_count += 1

            cap.release()
            print(f"    OCR完成: 共检测 {check_count} 帧")
            result['all_text'] = '\n'.join(detected_texts)
            result['success'] = True

        except Exception as e:
            result['error'] = str(e)

        # 保存到临时文件以便查看
        ocr_file = self.output_dir / "debug_ocr.txt"
        with open(ocr_file, 'w', encoding='utf-8') as f:
            f.write(result.get('all_text', '无内容'))
        print(f"\n  [OCR 识别结果已保存到：{ocr_file}]\n")

        return result

    def detect(self, video_path: Path, keywords: List[str], enable_speech: bool = True, enable_ocr: bool = True) -> Dict:
        """
        综合检测（语音 + OCR）

        Args:
            video_path: 视频路径
            keywords: 关键词列表
            enable_speech: 是否启用语音检测
            enable_ocr: 是否启用OCR检测

        Returns:
            {'speech': {...}, 'ocr': {...}}
        """
        result = {
            'video': str(video_path),
            'speech': {},
            'ocr': {}
        }

        # 语音检测
        if enable_speech and WHISPER_AVAILABLE:
            print(f"  [语音检测] 正在处理...")
            result['speech'] = self.detect_speech(video_path, keywords)
        elif enable_speech and not WHISPER_AVAILABLE:
            print(f"  [语音检测] 跳过 (未安装 faster-whisper)")

        # OCR 检测
        if enable_ocr and OCR_AVAILABLE:
            print(f"  [字幕检测] 正在处理...")
            result['ocr'] = self.detect_ocr(video_path, keywords)
        elif enable_ocr and not OCR_AVAILABLE:
            print(f"  [字幕检测] 跳过 (未安装 easyocr)")

        return result


class DouyinDownloader:
    """抖音视频下载器"""

    def __init__(self, output_dir: str = "./downloads", headless: bool = True):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.headless = headless
        self.browser: Optional[Browser] = None
        self.context = None
        self.playwright = None

    async def init_browser(self):
        """初始化浏览器"""
        self.playwright = await async_playwright().start()
        self.browser = await self.playwright.chromium.launch(headless=self.headless)
        # 创建一个持久化上下文用于请求
        self.context = await self.browser.new_context()

    async def close_browser(self):
        """关闭浏览器"""
        if self.context:
            await self.context.close()
        if self.browser:
            await self.browser.close()
        if self.playwright:
            await self.playwright.stop()

    def parse_url(self, url: str) -> Optional[str]:
        """从 URL 中提取视频 ID"""
        # 标准视频链接
        match = re.search(r'/video/(\d+)', url)
        if match:
            return match.group(1)

        # note 链接
        match = re.search(r'/note/(\d+)', url)
        if match:
            return match.group(1)

        # 搜索页面/精选页面链接 (modal_id 参数)
        match = re.search(r'modal_id=(\d+)', url)
        if match:
            return match.group(1)

        return None

    async def get_video_info_and_download(self, url: str, output_path: Path) -> tuple:
        """获取视频信息并下载完整视频"""
        video_info = {}
        api_captured = asyncio.Event()

        async def handle_response(response):
            nonlocal video_info
            if 'aweme/detail' in response.url or 'aweme/post' in response.url:
                try:
                    data = await response.json()
                    if 'aweme_detail' in data:
                        video_info = data['aweme_detail']
                        api_captured.set()
                    elif 'aweme_list' in data and data['aweme_list']:
                        video_info = data['aweme_list'][0]
                        api_captured.set()
                except Exception:
                    pass

        page = await self.context.new_page()

        try:
            page.on('response', handle_response)

            # 转换 URL
            modal_match = re.search(r'modal_id=(\d+)', url)
            if modal_match and '/video/' not in url:
                video_id = modal_match.group(1)
                url = f"https://www.douyin.com/video/{video_id}"

            await page.goto(url, wait_until='domcontentloaded', timeout=60000)

            # 等待 API 响应
            try:
                await asyncio.wait_for(api_captured.wait(), timeout=10)
            except asyncio.TimeoutError:
                video_info = await self.extract_from_page(page)

            if not video_info:
                return None, False

            # 提取下载 URL（优先使用 download_addr）
            video_url = self.extract_video_url(video_info)
            if not video_url:
                return video_info, False

            await page.close()

            # 使用 context.request 直接下载（携带 cookies）
            print(f"  正在下载完整视频...")
            try:
                # 设置请求头
                response = await self.context.request.get(
                    video_url,
                    headers={'Referer': url},
                    timeout=300000  # 5分钟超时
                )

                if response.ok:
                    body = await response.body()
                    if len(body) > 100000:  # 至少 100KB
                        with open(output_path, 'wb') as f:
                            f.write(body)
                        return video_info, True
            except Exception as e:
                print(f"  [DEBUG] Direct download failed: {e}")

            # 备用方案：页面拦截
            download_page = await self.context.new_page()
            video_data = None
            data_captured = asyncio.Event()

            async def handle_download_response(response):
                nonlocal video_data
                url_lower = response.url.lower()
                if 'douyinvod' in url_lower or 'bytedance' in url_lower:
                    try:
                        body = await response.body()
                        if len(body) > 100000:
                            video_data = body
                            data_captured.set()
                    except Exception:
                        pass

            download_page.on('response', handle_download_response)

            try:
                await download_page.set_extra_http_headers({'Referer': url})
                # 使用 domcontentloaded 而不是 networkidle
                await download_page.goto(video_url, wait_until='domcontentloaded', timeout=60000)

                # 等待视频数据
                try:
                    await asyncio.wait_for(data_captured.wait(), timeout=60)
                except asyncio.TimeoutError:
                    pass

            finally:
                await download_page.close()

            if video_data:
                with open(output_path, 'wb') as f:
                    f.write(video_data)

            return video_info, video_data is not None

        except Exception as e:
            print(f"  [DEBUG] Error: {e}")
            return None, False

    async def get_video_info(self, url: str) -> Optional[dict]:
        """使用浏览器访问页面，拦截 API 响应获取视频详情"""
        video_info = {}
        api_captured = asyncio.Event()

        async def handle_response(response):
            nonlocal video_info
            if 'aweme/detail' in response.url or 'aweme/post' in response.url:
                try:
                    data = await response.json()
                    if 'aweme_detail' in data:
                        video_info = data['aweme_detail']
                        api_captured.set()
                    elif 'aweme_list' in data and data['aweme_list']:
                        video_info = data['aweme_list'][0]
                        api_captured.set()
                except Exception:
                    pass

        page = await self.context.new_page()

        try:
            page.on('response', handle_response)

            # 如果是搜索页面链接，转换为标准视频链接
            modal_match = re.search(r'modal_id=(\d+)', url)
            if modal_match and '/video/' not in url:
                video_id = modal_match.group(1)
                url = f"https://www.douyin.com/video/{video_id}"

            await page.goto(url, wait_until='domcontentloaded', timeout=60000)

            try:
                await asyncio.wait_for(api_captured.wait(), timeout=10)
            except asyncio.TimeoutError:
                video_info = await self.extract_from_page(page)

            return video_info if video_info else None

        except Exception as e:
            print(f"获取视频信息失败: {e}")
            return None
        finally:
            await page.close()

    async def extract_from_page(self, page: Page) -> Optional[dict]:
        """从页面中提取视频信息"""
        try:
            content = await page.content()
            match = re.search(r'RENDER_DATA\s*=\s*({.*?})\s*</script>', content, re.DOTALL)
            if match:
                import html
                data_str = html.unescape(match.group(1))
                data = json.loads(data_str)
                if 'aweme' in data and 'detail' in data['aweme']:
                    return data['aweme']['detail']
                if 'aweme_detail' in data:
                    return data['aweme_detail']
            return None
        except Exception:
            return None

    def extract_video_url(self, video_info: dict) -> Optional[str]:
        """从视频信息中提取无水印视频 URL（优先使用 play_addr）"""
        try:
            # 注意：download_addr 现在返回带水印视频，不能使用！
            # 必须使用 play_addr 才能获取无水印视频
            paths = [
                ('video', 'play_addr', 'url_list'),
                ('video', 'play_addr_h264', 'url_list'),
                ('video', 'play_addr_265', 'url_list'),
            ]

            for path in paths:
                data = video_info
                for key in path:
                    if isinstance(data, dict) and key in data:
                        data = data[key]
                    else:
                        data = None
                        break

                if isinstance(data, list) and data:
                    url = data[0]
                    # 移除水印参数
                    url = re.sub(r'&watermark=1', '', url)
                    url = re.sub(r'watermark=1&', '', url)
                    return url

            return None
        except Exception:
            return None

    def get_video_title(self, video_info: dict) -> str:
        """获取视频标题"""
        try:
            desc = video_info.get('desc', '') or video_info.get('title', '未命名')
            # 清理文件名中的非法字符（包括换行符）
            desc = re.sub(r'[<>:"/\\|?*\r\n]', '', desc)
            # 清理 emoji 和其他特殊 Unicode 字符（Windows 文件名不支持）
            desc = re.sub(r'[\U0001F300-\U0001F9FF]', '', desc)  # emoji 范围
            desc = desc.strip()
            return desc[:100] if len(desc) > 100 else desc
        except Exception:
            return '未命名'

    def get_author_name(self, video_info: dict) -> str:
        """获取作者名称"""
        try:
            author = video_info.get('author', {})
            name = author.get('nickname', '') or author.get('unique_id', '') or '未知作者'
            name = re.sub(r'[<>:"/\\|?*]', '', name)
            return name
        except Exception:
            return '未知作者'

    def get_video_id(self, video_info: dict) -> str:
        """获取视频 ID"""
        return video_info.get('aweme_id', '') or video_info.get('video_id', 'unknown')

    async def download_video(self, video_url: str, output_path: Path) -> bool:
        """下载视频 - 使用浏览器上下文直接请求以携带正确的 cookies"""
        try:
            # 方法1: 使用浏览器的 request API 直接下载（携带 cookies）
            try:
                if self.context:
                    response = await self.context.request.get(video_url, timeout=60000)
                    if response.ok:
                        body = await response.body()
                        if len(body) > 10000:
                            with open(output_path, 'wb') as f:
                                f.write(body)
                            return True
            except Exception as e:
                print(f"  [DEBUG] Direct request failed: {e}")

            # 方法2: 通过页面访问并拦截响应
            page = await self.context.new_page()
            try:
                video_data = None
                data_captured = asyncio.Event()

                async def handle_response(response):
                    nonlocal video_data
                    content_type = response.headers.get('content-type', '')
                    url = response.url.lower()
                    # 扩展匹配条件
                    if ('video' in content_type or
                        'mp4' in url or
                        'douyinvod' in url or
                        'bytedance' in url or
                        'mpegurl' in content_type or
                        response.status == 206):  # Partial Content
                        try:
                            body = await response.body()
                            if len(body) > 10000:
                                video_data = body
                                data_captured.set()
                        except Exception:
                            pass

                page.on('response', handle_response)

                # 访问视频 URL
                await page.goto(video_url, wait_until='networkidle', timeout=60000)

                # 等待视频数据
                try:
                    await asyncio.wait_for(data_captured.wait(), timeout=10)
                except asyncio.TimeoutError:
                    pass

                if video_data:
                    with open(output_path, 'wb') as f:
                        f.write(video_data)
                    return True

                # 方法3: 尝试从页面获取 video 标签
                video_src = await page.evaluate('''
                    () => {
                        const video = document.querySelector('video');
                        return video ? (video.src || video.currentSrc) : null;
                    }
                ''')

                if video_src and video_src != video_url:
                    print(f"  [DEBUG] Found video src: {video_src[:80]}...")
                    # 递归尝试下载
                    return await self.download_video(video_src, output_path)

                return False

            finally:
                await page.close()

        except Exception as e:
            print(f"  [DEBUG] Download error: {e}")
            return False

    async def download_one(self, url: str, filename_format: str = "{title}",
                          keywords: List[str] = None, enable_speech: bool = True, enable_ocr: bool = True,
                          reporter: DetectionReporter = None) -> Optional[Path]:
        """下载单个视频"""
        print(f"正在处理: {url}")

        # 转换 URL（用于生成文件名）
        modal_match = re.search(r'modal_id=(\d+)', url)
        if modal_match and '/video/' not in url:
            video_id_for_path = modal_match.group(1)
            converted_url = f"https://www.douyin.com/video/{video_id_for_path}"
        else:
            converted_url = url

        # 生成临时输出路径
        temp_output = self.output_dir / f"temp_{datetime.now().strftime('%Y%m%d_%H%M%S')}.mp4"

        # 在同一页面获取信息并下载
        video_info, download_success = await self.get_video_info_and_download(url, temp_output)

        if not video_info:
            print(f"  [X] 无法获取视频信息")
            return None

        if not download_success:
            print(f"  [X] 下载失败")
            return None

        # 生成正式文件名
        title = self.get_video_title(video_info)
        author = self.get_author_name(video_info)
        video_id = self.get_video_id(video_info)

        filename = filename_format.format(title=title, author=author, id=video_id)
        filename = re.sub(r'[<>:"/\\|?*]', '', filename)
        output_path = self.output_dir / f"{filename}.mp4"

        # 检查是否已存在
        if output_path.exists():
            print(f"  [OK] 文件已存在: {output_path.name}")
            temp_output.unlink(missing_ok=True)
        else:
            # 重命名临时文件
            temp_output.rename(output_path)
            print(f"  [OK] 下载完成: {title[:50]}...")

        # 关键词检测
        if keywords and output_path.exists():
            detector = KeywordDetector(model_size=MODEL_SIZE, output_dir=self.output_dir)
            print(f"  正在检测关键词: {', '.join(keywords)}")

            detection_result = detector.detect(output_path, keywords, enable_speech, enable_ocr)

            speech_kw = detection_result['speech'].get('keywords_found', {})
            ocr_kw = detection_result['ocr'].get('keywords_found', {})
            full_speech = detection_result['speech'].get('transcript', '')
            full_ocr = detection_result['ocr'].get('all_text', '')

            # 单条检测时输出完整内容（保存到文件，避免终端编码问题）
            if not reporter:
                if full_speech:
                    speech_file = self.output_dir / "debug_speech.txt"
                    with open(speech_file, 'w', encoding='utf-8') as f:
                        f.write(full_speech)
                    print(f"\n  [语音识别结果已保存到：{speech_file}]\n")
                if full_ocr:
                    ocr_file = self.output_dir / "debug_ocr.txt"
                    with open(ocr_file, 'w', encoding='utf-8') as f:
                        f.write(full_ocr)
                    print(f"\n  [OCR 识别结果已保存到：{ocr_file}]\n")
            # 输出到报告文件
            if reporter:
                reporter.add_result(output_path.name, url, speech_kw, ocr_kw, full_speech, full_ocr)

            # 控制台简要输出
            if speech_kw or ocr_kw:
                print(f"  [!] 发现关键词!")
                if speech_kw:
                    for kw, occs in speech_kw.items():
                        print(f"      语音 \"{kw}\": {len(occs)} 次")
                        for occ in occs:
                            print(f"          - {occ['timestamp']}: {occ['text'][:50]}")
                if ocr_kw:
                    for kw, occs in ocr_kw.items():
                        print(f"      字幕 \"{kw}\": {len(occs)} 次")
                        for occ in occs:
                            print(f"          - {occ['timestamp']}: {occ['text'][:50]}")
            else:
                print(f"  [OK] 未发现关键词")

        return output_path

    async def download_from_excel(self, excel_path: str, filename_format: str = "{title}",
                                  keywords: List[str] = None, enable_speech: bool = True, enable_ocr: bool = True) -> list:
        """从 Excel 文件批量下载视频"""
        if openpyxl is None:
            print("请先安装 openpyxl: pip install openpyxl")
            return []

        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active

        urls = []
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0]:
                urls.append(str(row[0]).strip())

        print(f"从 Excel 读取到 {len(urls)} 个链接")

        # 创建报告器（Excel 模式）
        reporter = DetectionReporter(self.output_dir, excel_mode=True) if keywords else None

        # 存储检测结果
        detection_results = []

        for i, url in enumerate(urls, 1):
            print(f"\n[{i}/{len(urls)}]")
            result = await self.download_one(url, filename_format, keywords, enable_speech, enable_ocr, reporter)

            # 收集检测结果
            if reporter and reporter.results:
                last_result = reporter.results[-1]
                detection_results.append({
                    'url': url,
                    'success': result is not None,
                    'path': str(result) if result else None,
                    'speech_keywords': last_result.get('speech_keywords', {}),
                    'ocr_keywords': last_result.get('ocr_keywords', {}),
                    'full_speech': last_result.get('full_speech', ''),
                    'full_ocr': last_result.get('full_ocr', '')
                })
            else:
                detection_results.append({
                    'url': url,
                    'success': result is not None,
                    'path': str(result) if result else None,
                    'speech_keywords': {},
                    'ocr_keywords': {},
                    'full_speech': '',
                    'full_ocr': ''
                })

        # Excel 模式：写入结果到 Excel
        if reporter and reporter.excel_mode:
            self._write_results_to_excel(excel_path, detection_results)
            print(f"\n[OK] 检测结果已写入 Excel: {excel_path}")
        elif reporter:
            reporter.write_summary()

        return detection_results

    def _write_results_to_excel(self, excel_path: str, results: list):
        """将检测结果写入 Excel"""
        if openpyxl is None:
            print("[ERROR] openpyxl 未安装")
            return

        try:
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active

            # 在第 2 列添加标题（如果不存在）
            if ws['B1'] != '检测结果':
                ws['B1'] = '检测结果'

            for idx, result in enumerate(results, start=2):
                url = result['url']
                speech_kw = result.get('speech_keywords', {})
                ocr_kw = result.get('ocr_keywords', {})

                # 构建检测结果文本
                detection_text = ""

                # 语音关键词
                if speech_kw:
                    for kw, occurrences in speech_kw.items():
                        for occ in occurrences:
                            detection_text += f"[语音] {kw} @ {occ['timestamp']}: {occ['text'][:30]}\n"

                # OCR 关键词
                if ocr_kw:
                    for kw, occurrences in ocr_kw.items():
                        for occ in occurrences:
                            detection_text += f"[字幕] {kw} @ {occ['timestamp']}: {occ['text'][:30]}\n"

                # 如果没有检测到关键词
                if not detection_text:
                    detection_text = "未发现关键词"

                # 写入 Excel（第 2 列）
                ws[f'B{idx}'] = detection_text

            wb.save(excel_path)
            print(f"[OK] 检测结果已写入 Excel: {excel_path}")
        except PermissionError:
            # Excel 文件被占用，保存到备份文件
            backup_path = excel_path.rsplit('.', 1)[0] + "_results.xlsx"
            self._write_results_to_backup_excel(excel_path, backup_path, results)
        except Exception as e:
            print(f"[WARN] 无法直接写入 Excel: {e}")
            # 保存到备份文件
            backup_path = excel_path.rsplit('.', 1)[0] + "_results.xlsx"
            self._write_results_to_backup_excel(excel_path, backup_path, results)

    def _write_results_to_backup_excel(self, original_path: str, backup_path: str, results: list):
        """将检测结果写入备份 Excel 文件"""
        if openpyxl is None:
            print("[ERROR] openpyxl 未安装")
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "检测结果"

        # 写入标题
        ws['A1'] = '原始文件'
        ws['B1'] = 'URL'
        ws['C1'] = '检测结果'
        ws['D1'] = '语音关键词'
        ws['E1'] = '字幕关键词'

        # 写入数据
        for idx, result in enumerate(results, start=2):
            ws[f'A{idx}'] = original_path
            ws[f'B{idx}'] = result['url']

            # 构建检测结果文本
            detection_text = ""
            speech_kw = result.get('speech_keywords', {})
            ocr_kw = result.get('ocr_keywords', {})

            if speech_kw:
                for kw, occurrences in speech_kw.items():
                    for occ in occurrences:
                        detection_text += f"[语音] {kw} @ {occ['timestamp']}: {occ['text'][:30]}\n"

            if ocr_kw:
                for kw, occurrences in ocr_kw.items():
                    for occ in occurrences:
                        detection_text += f"[字幕] {kw} @ {occ['timestamp']}: {occ['text'][:30]}\n"

            if not detection_text:
                detection_text = "未发现关键词"

            ws[f'C{idx}'] = detection_text

            # 单独列出关键词
            all_keywords = set(speech_kw.keys()) | set(ocr_kw.keys())
            ws[f'D{idx}'] = ', '.join(all_keywords) if all_keywords else '-'

        wb.save(backup_path)
        print(f"[OK] 检测结果已保存到备份文件：{backup_path}")


async def main():
    parser = argparse.ArgumentParser(
        description='抖音视频无水印下载器 v1.2.4',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
示例:
  # 下载单个视频
  python douyin_downloader.py "https://www.douyin.com/video/7627089262673154289"

  # 从 Excel 批量下载
  python douyin_downloader.py -e videos.xlsx

  # 下载并检测关键词（语音+字幕）
  python douyin_downloader.py "URL" --detect "抖音"

  # 仅检测语音（禁用 OCR）
  python douyin_downloader.py "URL" --detect "抖音" --no-ocr

  # 仅检测字幕（禁用语音）
  python douyin_downloader.py "URL" --detect "抖音" --no-speech

  # 批量检测
  python douyin_downloader.py -e videos.xlsx --detect "抖音,广告"
        '''
    )

    parser.add_argument('url', nargs='?', help='视频链接')
    parser.add_argument('-e', '--excel', help='Excel 文件路径')
    parser.add_argument('-o', '--output', default='./downloads', help='输出目录')
    parser.add_argument('-f', '--format', default='{title}', help='文件名格式')
    parser.add_argument('--detect', help='检测关键词，多个用逗号分隔')
    parser.add_argument('--no-ocr', action='store_true', help='禁用字幕 OCR 检测')
    parser.add_argument('--no-speech', action='store_true', help='禁用语音检测')
    parser.add_argument('--model', default='base', help='语音识别模型')
    parser.add_argument('--head', action='store_true', help='显示浏览器窗口')

    args = parser.parse_args()

    if not args.url and not args.excel:
        parser.print_help()
        sys.exit(1)

    keywords = None
    enable_speech = not args.no_speech
    enable_ocr = not args.no_ocr

    if args.detect:
        keywords = [kw.strip() for kw in args.detect.split(',') if kw.strip()]
        if keywords:
            # 设置全局模型大小
            MODEL_SIZE = args.model
            if not WHISPER_AVAILABLE:
                print("[WARN] 请安装 faster-whisper: pip install faster-whisper")
            if not OCR_AVAILABLE and not args.no_ocr:
                print("[WARN] 请安装 OCR 依赖: pip install easyocr opencv-python")

    downloader = DouyinDownloader(output_dir=args.output, headless=not args.head)

    try:
        await downloader.init_browser()

        if args.excel:
            results = await downloader.download_from_excel(
                args.excel, args.format, keywords, enable_speech, enable_ocr
            )
            success_count = sum(1 for r in results if r['success'])
            print(f"\n{'='*50}")
            print(f"批量下载完成: {success_count}/{len(results)} 成功")
        else:
            # 单条检测时不创建 reporter，直接输出到控制台
            await downloader.download_one(
                args.url, args.format, keywords, enable_speech, enable_ocr, None
            )

    finally:
        await downloader.close_browser()


if __name__ == '__main__':
    asyncio.run(main())
